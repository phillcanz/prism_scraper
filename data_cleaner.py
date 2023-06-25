import os
import sys
import pandas as pd
import openpyxl

class DataCleaner:
    def create_excel(self):
        with open(os.path.join(sys.path[0], "policy_data.json"), "r") as readfile:
            data_panda = pd.read_json(readfile).transpose()
            data_panda.index.name = 'POLICIES'
            df_gen_info = pd.json_normalize(data_panda.loc[:,'GENERAL_INFORMATION']).set_index(data_panda.index)
            df_pay_info = pd.json_normalize(data_panda.loc[:,'PAYMENT_INFORMATION']).set_index(data_panda.index)
            df_pln_info = pd.json_normalize(data_panda.loc[:,'PLAN_DETAILS']).set_index(data_panda.index)
            df_fnd_info = pd.json_normalize(data_panda.loc[:,'FUND_DETAILS']).set_index(data_panda.index)
            df_ben_info = pd.json_normalize(data_panda.loc[:,'BENEFICIARY_DETAILS']).set_index(data_panda.index)
            

            df_pln_info.columns = df_pln_info.columns.astype(str).str.partition('.', expand=True).droplevel(level=1)
            df_fnd_info.columns = df_fnd_info.columns.astype(str).str.partition('.', expand=True).droplevel(level=1)
            df_ben_info.columns = df_ben_info.columns.astype(str).str.partition('.', expand=True).droplevel(level=1)
         
            df_pln_info.drop('COMPONENT_DESCRIPTION', axis=1, level=1, inplace=True)
            df_fnd_info.drop('FUND_TYPE', axis=1, level=1, inplace=True)

            with pd.ExcelWriter(os.path.join(sys.path[0], "policy_data.xlsx")) as writer:
                df_gen_info.to_excel(writer, sheet_name='GENERAL_INFORMATION')
                df_pay_info.to_excel(writer, sheet_name='PAYMENT_INFORMATION')
                df_pln_info.to_excel(writer, sheet_name='PLAN_DETAILS',)
                df_fnd_info.to_excel(writer, sheet_name='FUND_DETAILS')
                df_ben_info.to_excel(writer, sheet_name='BENEFICIARY_DETAILS')

            wb = openpyxl.load_workbook(os.path.join(sys.path[0], "policy_data.xlsx"))
            ws_plan = wb['PLAN_DETAILS']
            ws_plan['A1'] = ws_plan['A3'].value
            ws_plan.merge_cells('A1:A3')
            ws_fund = wb['FUND_DETAILS']
            ws_fund['A1'] = ws_fund['A3'].value
            ws_fund.merge_cells('A1:A3')
            ws_beni = wb['BENEFICIARY_DETAILS']
            ws_beni['A1'] = ws_beni['A3'].value
            ws_beni.merge_cells('A1:A3')

            wb.save((os.path.join(sys.path[0], "policy_data.xlsx")))
            # df_pln_info.to_excel(excel_writer=os.path.join(
            #     sys.path[0], "sample_info.xlsx"), merge_cells=True)

    def clean_excel(self):
        pass
if __name__ == "__main__":
    dc = DataCleaner()
    dc.create_excel()